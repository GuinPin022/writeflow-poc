from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

ACCENT = "2B579A"
LIGHT = "E8F0FE"
GREY = "F3F4F6"
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name="Arial", bold=True, size=14, color=ACCENT)
SUB_FONT = Font(name="Arial", bold=True, size=11, color=ACCENT)
BASE_FONT = Font(name="Arial", size=10)
BOLD = Font(name="Arial", size=10, bold=True)
header_fill = PatternFill("solid", fgColor=ACCENT)
light_fill = PatternFill("solid", fgColor=LIGHT)
grey_fill = PatternFill("solid", fgColor=GREY)
thin = Side(style="thin", color="D0D0D0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="center")

wb = Workbook()

# ---------------- Sheet 1: Instructions ----------------
ws = wb.active
ws.title = "Instructions"
ws.sheet_view.showGridLines = False
ws["B2"] = "WriteFlow — Journal de tests du POC tracking"
ws["B2"].font = TITLE_FONT
lines = [
    ("Objectif", "Consigner chaque test du POC et comparer la mesure de l'add-in a la verite terrain (mots reellement tapes/colles/supprimes, comptes a la main)."),
    ("Onglet 'Releves'", "Une ligne par test. Remplir les colonnes bleu clair (saisie). Les colonnes grises (erreur %, verdict) se calculent seules."),
    ("Verite terrain", "Avant chaque test, noter le nombre de mots reellement produits. Puis lancer le scenario dans Word avec l'add-in, et reporter les valeurs des exports CSV/JSON."),
    ("Onglet 'Synthese'", "Agrege les resultats et propose un verdict global GO / GO conditionnel / NO-GO selon les seuils du protocole."),
    ("Seuils (rappel)", "Erreur tapee < 2% = GO, 2-8% = conditionnel, > 8% = NO-GO. Latence < 200 ms = GO, > 500 ms = NO-GO. Detection collage > 95% = GO."),
    ("Conseil", "Rejouer le MEME scenario sur Desktop Windows, Desktop Mac et Web pour tester la coherence cross-plateforme."),
]
r = 4
for k, v in lines:
    ws[f"B{r}"] = k
    ws[f"B{r}"].font = BOLD
    ws[f"C{r}"] = v
    ws[f"C{r}"].font = BASE_FONT
    ws[f"C{r}"].alignment = wrap
    ws.row_dimensions[r].height = 42
    r += 1
ws.column_dimensions["A"].width = 2
ws.column_dimensions["B"].width = 20
ws.column_dimensions["C"].width = 95

# ---------------- Sheet 2: Releves ----------------
rl = wb.create_sheet("Releves")
rl.sheet_view.showGridLines = False
headers = [
    ("ID", 6), ("Date", 12), ("Plateforme", 18), ("Scenario", 28),
    ("VT mots tapes", 12), ("VT mots colles", 12), ("VT mots supprimes", 13),
    ("Mesure tapee", 12), ("Mesure colle", 12), ("Mesure supprime", 13),
    ("Erreur tapee %", 13), ("Latence max (ms)", 13), ("WordApi 1.5 ?", 12),
    ("Collage attendu ?", 13), ("Collage detecte ?", 13), ("Verdict ligne", 14), ("Notes", 40),
]
rl["A1"] = "Releves de test — POC tracking"
rl["A1"].font = TITLE_FONT
hr = 3
for c, (name, w) in enumerate(headers, start=1):
    cell = rl.cell(row=hr, column=c, value=name)
    cell.font = HEADER_FONT
    cell.fill = header_fill
    cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    cell.border = border
    rl.column_dimensions[cell.column_letter].width = w
rl.row_dimensions[hr].height = 30

# input (blue-clair) columns vs computed (grey)
input_cols = {1,2,3,4,5,6,7,8,9,10,13,14,15,17}
computed_cols = {11,16}

first = hr + 1
last = first + 29  # 30 rows
for row in range(first, last + 1):
    for c in range(1, len(headers) + 1):
        cell = rl.cell(row=row, column=c)
        cell.font = BASE_FONT
        cell.border = border
        if c in computed_cols:
            cell.fill = grey_fill
        elif c in input_cols:
            cell.fill = light_fill
        if c in (5,6,7,8,9,10,12):
            cell.alignment = center
    # Erreur tapee % (col K=11): compare mesure tapee (H=8) a VT tapes (E=5)
    rl.cell(row=row, column=11,
            value=f'=IF($E{row}="","",IF($E{row}=0,IF($H{row}=0,0,1),ABS($H{row}-$E{row})/$E{row}))')
    rl.cell(row=row, column=11).number_format = "0.0%"
    # Verdict ligne (col P=16)
    rl.cell(row=row, column=16,
            value=(f'=IF($E{row}="","",'
                   f'IF(OR($K{row}>0.08,$L{row}>500),"NO-GO",'
                   f'IF(AND($K{row}<=0.02,$L{row}<=200,OR($N{row}="Non",AND($N{row}="Oui",$O{row}="Oui"))),"GO",'
                   f'"Conditionnel")))'))
    rl.cell(row=row, column=16).alignment = center
    rl.cell(row=row, column=16).font = BOLD

# Data validations
dv_plat = DataValidation(type="list", formula1='"Desktop Windows,Desktop Mac,Web,iPad"', allow_blank=True)
dv_scen = DataValidation(type="list", formula1='"Frappe continue,Coller un bloc,Couper-supprimer,Reecriture/revision,Dictee vocale,Autocorrection,Annuler-Retablir,Rechercher-remplacer,Deplacer un bloc,Notes de bas de page,Document long 80-100k,Track changes,Co-edition,Inactivite 15min,Hors-ligne"', allow_blank=True)
dv_oui = DataValidation(type="list", formula1='"Oui,Non"', allow_blank=True)
for dv in (dv_plat, dv_scen, dv_oui):
    rl.add_data_validation(dv)
dv_plat.add(f"C{first}:C{last}")
dv_scen.add(f"D{first}:D{last}")
dv_oui.add(f"M{first}:M{last}")
dv_oui.add(f"N{first}:N{last}")
dv_oui.add(f"O{first}:O{last}")

# Example row
ex = [1, "2026-06-20", "Desktop Windows", "Coller un bloc", 0, 2000, 0, 0, 1980, 0, None, 95, "Oui", "Oui", "Oui", None, "Exemple : collage de 2000 mots, bien classe en colle."]
for c, v in enumerate(ex, start=1):
    if v is not None:
        rl.cell(row=first, column=c, value=v)
rl.freeze_panes = "A4"

# ---------------- Sheet 3: Synthese ----------------
sy = wb.create_sheet("Synthese")
sy.sheet_view.showGridLines = False
sy["B2"] = "Synthese & decision"
sy["B2"].font = TITLE_FONT
rng_k = f"Releves!K{first}:K{last}"
rng_l = f"Releves!L{first}:L{last}"
rng_p = f"Releves!P{first}:P{last}"
rng_n = f"Releves!N{first}:N{last}"
rng_o = f"Releves!O{first}:O{last}"
rng_a = f"Releves!A{first}:A{last}"

metrics = [
    ("Tests realises", f"=COUNTA({rng_a})", "0"),
    ("Erreur tapee moyenne", f'=IFERROR(AVERAGE({rng_k}),"-")', "0.0%"),
    ("Erreur tapee max", f'=IFERROR(MAX({rng_k}),"-")', "0.0%"),
    ("Latence moyenne (ms)", f'=IFERROR(AVERAGE({rng_l}),"-")', "0"),
    ("Latence max (ms)", f'=IFERROR(MAX({rng_l}),"-")', "0"),
    ("Taux detection collage", f'=IFERROR(COUNTIFS({rng_n},"Oui",{rng_o},"Oui")/COUNTIF({rng_n},"Oui"),"-")', "0.0%"),
    ("Tests GO", f'=COUNTIF({rng_p},"GO")', "0"),
    ("Tests Conditionnel", f'=COUNTIF({rng_p},"Conditionnel")', "0"),
    ("Tests NO-GO", f'=COUNTIF({rng_p},"NO-GO")', "0"),
]
sy["B4"] = "Indicateurs"
sy["B4"].font = SUB_FONT
r = 5
for label, formula, fmt in metrics:
    sy[f"B{r}"] = label
    sy[f"B{r}"].font = BASE_FONT
    sy[f"B{r}"].fill = grey_fill
    sy[f"B{r}"].border = border
    c = sy[f"C{r}"]
    c.value = formula
    c.font = BOLD
    c.number_format = fmt
    c.border = border
    c.alignment = center
    r += 1

# Verdict global
sy[f"B{r+1}"] = "Verdict global"
sy[f"B{r+1}"].font = SUB_FONT
vc = sy[f"B{r+2}"]
vc.value = (f'=IF(COUNTA({rng_a})=0,"Aucun test saisi",'
            f'IF(COUNTIF({rng_p},"NO-GO")>0,"NO-GO : au moins un test critique echoue, revoir avant de continuer",'
            f'IF(COUNTIF({rng_p},"Conditionnel")>0,"GO CONDITIONNEL : documenter les limites avant le MVP",'
            f'"GO : tracking fiable sur les tests realises")))')
vc.font = Font(name="Arial", bold=True, size=12, color=ACCENT)
vc.fill = light_fill
vc.alignment = Alignment(wrap_text=True, vertical="center")
sy.merge_cells(start_row=r+2, start_column=2, end_row=r+2, end_column=6)
sy.row_dimensions[r+2].height = 36

# Thresholds reminder
tr = r + 4
sy[f"B{tr}"] = "Seuils de decision (rappel du protocole)"
sy[f"B{tr}"].font = SUB_FONT
th = [
    ("Critere", "GO", "Conditionnel", "NO-GO"),
    ("Erreur de comptage", "< 2%", "2 - 8%", "> 8%"),
    ("Detection collage", "> 95%", "80 - 95%", "< 80%"),
    ("Latence (doc long)", "< 200 ms", "200 - 500 ms", "> 500 ms"),
    ("Coherence cross-plateforme", "identique", "ecarts mineurs", "1 plateforme KO"),
    ("Perte hors-ligne", "0", "rare", "frequente"),
]
for i, rowv in enumerate(th):
    for j, val in enumerate(rowv):
        cell = sy.cell(row=tr+1+i, column=2+j, value=val)
        cell.border = border
        if i == 0:
            cell.font = HEADER_FONT
            cell.fill = header_fill
            cell.alignment = center
        else:
            cell.font = BASE_FONT if j == 0 else Font(name="Arial", size=10, italic=(j>0))
            if j == 0:
                cell.fill = grey_fill
sy.column_dimensions["B"].width = 26
for col in ("C", "D", "E", "F"):
    sy.column_dimensions[col].width = 18

wb.save("/sessions/keen-wizardly-pasteur/mnt/outputs/WriteFlow-journal-tests-POC.xlsx")
print("saved")
